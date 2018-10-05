# update_produce.py
# This program will open a spreadsheet and update various produce prices

import openpyxl


def open_workbook():
    ''' This function will attempt to open an excel workbook and return it as a workbook object.
        Returns --- OpenPyXl workbook object
    '''
    try:
        workbook = openpyxl.load_workbook('produceSales.xlsx')
    except:
        print('Unable to open spreadsheet.')

    return workbook


def update_item(sheet):
    ''' This function will search the rows in the sheet object for items
        that need to have their prices updated and updates them.

        Arguments:
        sheet -- an openpyxl sheet object to search
    '''
    # The produce types and their price
    PRICE_UPDATES = {'Garlic': 3.07,
                     'Celery': 1.19,
                     'Lemon': 1.27}

    for row_num in range(2, sheet.max_row + 1):  # skip row 1 as it's header info
        produce_name = sheet.cell(row=row_num, column=1).value

        if produce_name in PRICE_UPDATES:
            sheet.cell(
                row=row_num, column=2).value = PRICE_UPDATES[produce_name]
            print(
                f'{produce_name} price updated to {PRICE_UPDATES[produce_name]}')


def main():
    ''' This is the main function of the program. '''

    wb = open_workbook()

    sheet = wb.active   # This will get the active sheet
    update_item(sheet)

    wb.save('updatedProduceSales.xlsx')


if __name__ == '__main__':
    main()
